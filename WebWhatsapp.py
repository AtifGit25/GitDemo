import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
service_obj = Service("C:\\Users\\Admin\\Desktop\\chromedriver.exe")
driver = webdriver.Chrome(service=service_obj)
driver.maximize_window()
wb = openpyxl.load_workbook('D:\\Whatsapp.xlsx')
sheet = wb.active

Message = sheet.cell(2,2)

for col_cells in sheet.iter_cols(min_col=1, max_col=1):
    for Num in col_cells:
        driver.get('https://wa.me/+91{}'.format(Num.value))
        driver.implicitly_wait(10)
        driver.find_element(By.LINK_TEXT, "Continue to Chat").click()
        driver.implicitly_wait(10)
        driver.find_element(By.LINK_TEXT, "use WhatsApp Web").click()
        driver.implicitly_wait(10)
        time.sleep(20)
        driver.implicitly_wait(20)
        driver.find_element(By.XPATH, "//div[@title='Type a message']").send_keys(Message.value)
        driver.find_element(By.XPATH, "//div[@class='_3HQNh _1Ae7k']/button/span[1]").click()
        time.sleep(5)
